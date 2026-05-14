#!/usr/bin/env python3
"""Hermes AI 快消版 - xlsx 解析器
用法: python3 parse_xlsx.py <action> <file1> [file2]

action: loss | reconcile | order | collection
"""

import sys, json, os

# 将 python-lib 加入 sys.path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LIB_DIR = os.path.join(SCRIPT_DIR, 'python-lib')
if os.path.isdir(LIB_DIR):
    sys.path.insert(0, LIB_DIR)

from openpyxl import load_workbook


def read_sheet(filepath, sheet_idx=0):
    """读取 xlsx 的指定 sheet，返回二维数组"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.worksheets[sheet_idx]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else '' for v in row])
    wb.close()
    return rows


def calc_loss(file1, file2):
    """货损计算"""
    data1 = read_sheet(file1)
    data2 = read_sheet(file2)
    
    return {
        'action': 'loss',
        'file1_rows': len(data1),
        'file2_rows': len(data2),
        'file1_cols': len(data1[0]) if data1 else 0,
        'file2_cols': len(data2[0]) if data2 else 0,
        'file1_header': data1[4] if len(data1) > 4 else [],
        'file2_header': data2[4] if len(data2) > 4 else [],
        'preview1': data1[:8],
        'preview2': data2[:8],
    }


def calc_reconcile(file1, file2):
    """对账 - 基础解析"""
    data1 = read_sheet(file1)
    data2 = read_sheet(file2)
    return {
        'action': 'reconcile',
        'file1_rows': len(data1),
        'file2_rows': len(data2),
        'file1_header': data1[4] if len(data1) > 4 else [],
        'file2_header': data2[0] if data2 else [],
        'preview1': data1[:8],
        'preview2': data2[:8],
    }


def calc_order(file1):
    """订单模板"""
    data = read_sheet(file1)
    return {
        'action': 'order',
        'rows': len(data),
        'header': data[0] if data else [],
        'preview': data[:5],
    }


def calc_collection(file1, file2):
    """催收看板"""
    data1 = read_sheet(file1)
    data2 = read_sheet(file2)
    return {
        'action': 'collection',
        'file1_rows': len(data1),
        'file2_rows': len(data2),
        'preview1': data1[:8],
        'preview2': data2[:8],
    }


def format_report(result):
    """将解析结果格式化为文本报告"""
    action = result['action']
    
    if action == 'loss':
        lines = ['📊 货损报告（已解析Excel，待详细计算）']
        lines.append(f'')
        lines.append(f'调入文件行数: {result["file1_rows"]}')
        lines.append(f'调出文件行数: {result["file2_rows"]}')
        lines.append(f'')
        lines.append(f'【调入表头】')
        lines.append(f'  {", ".join(str(x)[:10] for x in result["file1_header"])}')
        lines.append(f'')
        lines.append(f'【调出表头】')
        lines.append(f'  {", ".join(str(x)[:10] for x in result["file2_header"])}')
        lines.append(f'')
        lines.append('⚠️ 详细货损率计算需要配置具体列映射')
        return '\n'.join(lines)
    
    elif action == 'reconcile':
        lines = ['📊 银行流水对账（已解析Excel）']
        lines.append(f'')
        lines.append(f'舟谱流水行数: {result["file1_rows"]}')
        lines.append(f'银行流水行数: {result["file2_rows"]}')
        lines.append(f'')
        lines.append(f'【舟谱表头】')
        lines.append(f'  {", ".join(str(x)[:10] for x in result["file1_header"])}')
        lines.append(f'')
        lines.append(f'【银行表头】')
        lines.append(f'  {", ".join(str(x)[:10] for x in result["file2_header"])}')
        return '\n'.join(lines)
    
    elif action == 'order':
        lines = ['📋 订单导入模板']
        lines.append(f'行数: {result["rows"]}')
        lines.append(f'')
        lines.append(f'【表头】')
        lines.append(f'  {", ".join(str(x)[:10] for x in result["header"])}')
        return '\n'.join(lines)
    
    elif action == 'collection':
        lines = ['💰 催收看板（已解析Excel）']
        lines.append(f'')
        lines.append(f'应收款行数: {result["file1_rows"]}')
        lines.append(f'回款流水行数: {result["file2_rows"]}')
        return '\n'.join(lines)
    
    return json.dumps(result, ensure_ascii=False, indent=2)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({'error': '参数不足'}))
        sys.exit(1)
    
    action = sys.argv[1]
    
    try:
        if action == 'loss':
            result = calc_loss(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else sys.argv[2])
        elif action == 'reconcile':
            result = calc_reconcile(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else sys.argv[2])
        elif action == 'order':
            result = calc_order(sys.argv[2])
        elif action == 'collection':
            result = calc_collection(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else sys.argv[2])
        else:
            result = {'error': f'未知操作: {action}'}
        
        # 输出 JSON 报告
        report = format_report(result)
        print(json.dumps({'success': True, 'action': action, 'report': report, **result}, ensure_ascii=False))
        
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False))
        sys.exit(1)
